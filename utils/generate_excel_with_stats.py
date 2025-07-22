import sys
import os
import pandas as pd
import numpy as np
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font
from openpyxl.utils import get_column_letter

def round_or_div0(value):
    if pd.isna(value):
        return '#DIV/0!'
    return round(value, 3)

def generate_excel_from_csv(csv_path):
    # Load CSV
    df = pd.read_csv(csv_path)

    # Auto-detect base prefix from actual column names
    sample_block = "1k"
    sample_suffix = "Ans_best"
    matching_cols = [col for col in df.columns if f"_{sample_block}_{sample_suffix}" in col]
    if not matching_cols:
        raise ValueError("❌ Could not detect a valid column to infer prefix (e.g., *_1k_Ans_best)")
    first_match = matching_cols[0]
    base_prefix = first_match.rsplit(f"_{sample_block}_{sample_suffix}", 1)[0]
    print(f"✅ Detected base prefix: {base_prefix}")

    # Metadata
    methods = ['Ans', 'Classifier', 'MLP', 'BF', 'SA', 'GA']
    variants = ['best', '5th', '10th']
    blocks = [f"{i}k" for i in range(1, 21)]
    filtered_cols = [f"{m}_{v}" for m in methods for v in variants if not (m in ['SA', 'GA'] and v != 'best')]
    header = ['', '', *filtered_cols]

    # Create Excel workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "Results"
    ws.append(header)

    # Style header
    gray_fill = PatternFill(start_color='D9D9D9', end_color='D9D9D9', fill_type='solid')
    bold_font = Font(bold=True)
    for cell in ws[1]:
        cell.font = bold_font
        cell.fill = gray_fill

    # Output rows for each block (1k, 2k, ..., 20k)
    for idx, block in enumerate(blocks):
        prefix = f"{base_prefix}_{block}_"
        ordered_cols = [prefix + col for col in filtered_cols]
        available_cols = [col for col in ordered_cols if col in df.columns]

        block_df = df[available_cols].apply(pd.to_numeric, errors='coerce')
        block_df = block_df.reindex(columns=ordered_cols)  # Fill missing with NaN

        avg = block_df.mean(skipna=True)
        std = block_df.std(skipna=True, ddof=1)

        row_label = 100 + idx * 200
        ws.append([row_label, 'AVRG'] + list(map(round_or_div0, avg)))
        ws.append(['', 'STD'] + list(map(round_or_div0, std)))

    # Merge row labels in column A
    for i in range(2, 42, 2):
        ws.merge_cells(start_row=i, start_column=1, end_row=i+1, end_column=1)

    # Set column widths
    for i in range(1, len(header) + 1):
        ws.column_dimensions[get_column_letter(i)].width = 15

    # Save Excel file
    base, _ = os.path.splitext(csv_path)
    out_path = base + "_data.xlsx"
    wb.save(out_path)
    print(f"✅ Saved: {out_path}")

if __name__ == "__main__":
    if len(sys.argv) != 2:
        print("Usage: python generate_excel_with_stats.py <input.csv>")
        sys.exit(1)
    generate_excel_from_csv(sys.argv[1])
